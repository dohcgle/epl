
import openpyxl
import os

files = [
    '/home/ulugbek/Projects/epl/material/TURAYEV ISROIL BURIYEVICH.xlsx'
]

sheets_to_check = ['dalolatnoma', 'buyruq', 'monitoring_1', 'monitoring_2', 'monitoring_3', 'monitoring_4', 'grafik']

for f in files:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sheet_name in sheets_to_check:
            if sheet_name in wb.sheetnames:
                print(f"\n--- Reading sheet: {sheet_name} (First 20 rows) ---")
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    content = [str(c).strip() for c in row if c]
                    if content:
                        print(content)
            else:
                print(f"Sheet {sheet_name} not found.")
    except Exception as e:
        print(f"Error: {e}")
