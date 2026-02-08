
from django.test import TestCase, Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from .models import ProcessedDocument
import os
import openpyxl
import zipfile
import io

class DocumentOptionsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse('process_document')
        
        # Create a dummy excel file with specific sheets
        self.excel_path = 'test_doc_options.xlsx'
        wb = openpyxl.Workbook()
        # Rename default sheet
        ws1 = wb.active
        ws1.title = "Anketa"
        ws1['A1'] = "Anketa Data"
        
        ws2 = wb.create_sheet("Buyruq")
        ws2['A1'] = "Buyruq Data"
        
        ws3 = wb.create_sheet("Shartnoma")
        ws3['A1'] = "Shartnoma Data"
        
        ws4 = wb.create_sheet("Other")
        ws4['A1'] = "Other Data"
        
        wb.save(self.excel_path)
        
        with open(self.excel_path, 'rb') as f:
            self.file_content = f.read()

    def tearDown(self):
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        # Clean up processed docs
        for doc in ProcessedDocument.objects.all():
            if doc.original_file and os.path.exists(doc.original_file.path):
                os.remove(doc.original_file.path)
            if doc.processed_file and os.path.exists(doc.processed_file.path):
                os.remove(doc.processed_file.path)

    def test_no_selection(self):
        file = SimpleUploadedFile("test.xlsx", self.file_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        data = {'file': file} # No checkboxes
        
        response = self.client.post(self.url, data)
        self.assertEqual(response.status_code, 200)
        json_resp = response.json()
        self.assertIn("Hech qanday hujjat turi tanlanmadi", json_resp['message'])
        
        doc = ProcessedDocument.objects.last()
        self.assertFalse(bool(doc.processed_file))

    def test_single_selection(self):
        file = SimpleUploadedFile("test.xlsx", self.file_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        data = {
            'file': file,
            'doc_type_anketa': 'on'
        }
        
        response = self.client.post(self.url, data)
        self.assertEqual(response.status_code, 200)
        
        doc = ProcessedDocument.objects.last()
        self.assertTrue(bool(doc.processed_file))
        
        # Verify ZIP content
        with zipfile.ZipFile(doc.processed_file.path, 'r') as zf:
            names = zf.namelist()
            self.assertIn('Anketa.pdf', names)
            self.assertNotIn('Buyruq.pdf', names)
            self.assertNotIn('Shartnoma.pdf', names)
            self.assertNotIn('Other.pdf', names)

    def test_multiple_selection(self):
        file = SimpleUploadedFile("test.xlsx", self.file_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        data = {
            'file': file,
            'doc_type_anketa': 'on',
            'doc_type_buyruq': 'on'
        }
        
        response = self.client.post(self.url, data)
        self.assertEqual(response.status_code, 200)
        
        doc = ProcessedDocument.objects.last()
        with zipfile.ZipFile(doc.processed_file.path, 'r') as zf:
            names = zf.namelist()
            self.assertIn('Anketa.pdf', names)
            self.assertIn('Buyruq.pdf', names)
            self.assertNotIn('Shartnoma.pdf', names)

    def test_case_insensitive_match(self):
        # View passes "Shartnoma", sheet is "Shartnoma" -> verify
        # What if sheet was lowercase in Excel? 
        # Helper has case-insensitivity. Let's create a file with lowercase sheet.
        
        path = 'test_lower.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "shartnoma" # lowercase
        wb.save(path)
        
        with open(path, 'rb') as f:
            content = f.read()
        os.remove(path)
        
        file = SimpleUploadedFile("test_lower.xlsx", content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        data = {
            'file': file,
            'doc_type_shartnoma': 'on' # View passes "Shartnoma"
        }
        
        response = self.client.post(self.url, data)
        doc = ProcessedDocument.objects.last()
        with zipfile.ZipFile(doc.processed_file.path, 'r') as zf:
            # Filename in ZIP comes from sheet name "shartnoma"
            self.assertIn('shartnoma.pdf', zf.namelist())


