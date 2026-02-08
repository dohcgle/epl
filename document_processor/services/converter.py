import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border
from weasyprint import HTML, CSS
from django.template import Context, Template
import os

class ExcelToPdfConverter:
    def __init__(self, excel_path):
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)

    def convert_all(self, qr_generator_func, allowed_sheets=None):
        """
        Converts all sheets to PDF bytes.
        qr_generator_func: function that takes (sheet_name) and returns base64 QR code.
        allowed_sheets: list of sheet names (strings) to convert. If None or empty, converts none (or all? Logic says specific selection required).
                        However, usually 'None' implies default behavior. 
                        But our view passes explicit list.
        Returns a list of tuples: (sheet_name, pdf_bytes)
        """
        results = []
        
        # Normalize allowed_sheets for case-insensitive matching
        allowed_set = set(s.lower() for s in allowed_sheets) if allowed_sheets else set()

        for sheet_name in self.workbook.sheetnames:
            # Filter
            if allowed_sheets is not None:
                # Check if current sheet matches any allowed name
                if sheet_name.lower() not in allowed_set:
                    continue

            sheet = self.workbook[sheet_name]
            
            # Generate unique QR for this sheet
            qr_code_base64 = qr_generator_func(sheet_name)
            
            html_content = self._sheet_to_html(sheet, qr_code_base64)
            
            # Generate PDF using WeasyPrint
            css = CSS(string='''
                @page { size: A4 landscape; margin: 1cm; }
                body { font-family: "Times New Roman", Times, serif; }
                table { border-collapse: collapse; width: 100%; }
                td, th { padding: 4px; }
                .qr-code { 
                    position: fixed; 
                    bottom: 20px; 
                    right: 20px; 
                    width: 100px; 
                    height: 100px; 
                    z-index: 1000;
                }
            ''')
            
            try:
                pdf_bytes = HTML(string=html_content).write_pdf(stylesheets=[css])
                results.append((sheet_name, pdf_bytes))
            except Exception as e:
                print(f"Error converting sheet {sheet_name}: {e}")
                continue
            
        return results

    def _sheet_to_html(self, sheet, qr_code_base64):
        """
        Generates HTML from a single Excel sheet, preserving styles and fitting to A4.
        """
        # 1. Calculate Active Area (Bounding Box)
        # Check for Print Area first
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column
        
        use_print_area = False
        if sheet.print_area:
            try:
                # print_area can be a list "Sheet1!$A$1:$F$20" or just range "A1:F20"
                # OpenPyXL might return a list of ranges if multiple areas are selected.
                # We usually support only the first area for PDF generation simplicity.
                
                # Check if it's a list or string
                print_areas = sheet.print_area
                if isinstance(print_areas, list):
                    print_area = print_areas[0] # Take first
                else:
                    # It might be "Sheet1!$A$1:$F$20" string
                    print_area = str(print_areas).split(' ')[0] # split by space if multiple? usually comma

                # range_boundaries expects "A1:F20" w/o sheet name usually, but let's see.
                from openpyxl.utils import range_boundaries
                # Remove sheet name if present
                if '!' in print_area:
                    print_area = print_area.split('!')[1]
                
                min_col, min_row, max_col, max_row = range_boundaries(print_area)
                use_print_area = True
            except Exception as e:
                print(f"Error parsing print area '{sheet.print_area}': {e}. Falling back to auto-detect.")
                use_print_area = False

        if not use_print_area:
            # Refine max_row/max_col by checking for actual data (backwards)
            # Find last non-empty row
            found_data = False
            for r in range(max_row, min_row - 1, -1):
                empty = True
                for c in range(min_col, max_col + 1):
                    if sheet.cell(row=r, column=c).value is not None:
                        empty = False
                        break
                if not empty:
                    max_row = r
                    found_data = True
                    break
            
            if not found_data:
                # Empty sheet
                return "<html><body><p>Empty Sheet</p></body></html>"

            # Find last non-empty col
            for c in range(max_col, min_col - 1, -1):
                empty = True
                for r in range(min_row, max_row + 1):
                    if sheet.cell(row=r, column=c).value is not None:
                        empty = False
                        break
                if not empty:
                    max_col = c
                    break

        # 2. Calculate widths for Fit-to-Page
        # Standard A4 Landscape printable width approx 1000px-1100px (depending on DPI, weasyprint default is 96dpi -> 297mm = 1122px. Margins 1cm each side -> 20mm off -> 277mm ~= 1047px)
        # Let's sum huge excel column widths.
        # OpenPyXL width unit is num characters approx. 1 unit ~= 7px (roughly).
        
        total_excel_width = 0
        col_widths = [] # store percentage or px
        
        for i in range(min_col, max_col + 1):
            col_letter = get_column_letter(i)
            # sheet.column_dimensions[col_letter].width might be None (default)
            dims = sheet.column_dimensions.get(col_letter)
            w = dims.width if dims and dims.width else 8.43 # Default Excel width
            total_excel_width += w
            col_widths.append(w)

        # Heuristic: 1 Excel width unit ~= 7.5px.
        estimated_px_width = total_excel_width * 7.5 
        
        # Determine Page Orientation
        orientation = 'landscape' # Default
        if sheet.page_setup and sheet.page_setup.orientation:
            orientation = sheet.page_setup.orientation # 'portrait' or 'landscape'
        
        # A4 Landscape (297mm) minus margins (20mm) = 277mm ~= 1047px at 96dpi
        # A4 Portrait (210mm) minus margins (20mm) = 190mm ~= 718px at 96dpi
        
        if orientation == 'landscape':
            page_width_px = 1050
        else:
            page_width_px = 720
        
        scale_factor = 1.0
        
        if estimated_px_width > page_width_px:
            scale_factor = page_width_px / estimated_px_width
            # Limit min scale to avoid microscopic text
            if scale_factor < 0.3: scale_factor = 0.3
            
        # 3. Generate HTML
        html = ['<html><head><style>']
        
        # Add page CSS with dynamic transform
        # We transform the TABLE or BODY. 
        # transform-origin: top left;
        html.append(f'''
            @page {{ size: A4 {orientation}; margin: 1cm; }}
            body {{ 
                font-family: "Times New Roman", Times, serif; 
                margin: 0; padding: 0;
            }}
            .wrapper {{
                width: {estimated_px_width}px; /* Force container width */
                transform: scale({scale_factor});
                transform-origin: 0 0;
            }}
            table {{ border-collapse: collapse; table-layout: fixed; width: 100%; }}
            td, th {{ padding: 2px 4px; overflow: hidden; }}
            .qr-code {{ 
                position: fixed; 
                bottom: 20px; 
                right: 20px; 
                width: 100px; 
                height: 100px; 
                z-index: 1000;
            }}
        ''')
        html.append('</style></head><body>')
        
        # QR Code - Fix position? If we scale wrapper, QR might be affected if inside wrapper.
        # Put QR outside wrapper to keep it standard size/position on page?
        # Yes, QR should be relative to Page, not Content scale.
        html.append(f'<div class="wrapper">')
        html.append('<table>')
        
        # Colgroup for widths
        html.append('<colgroup>')
        for w in col_widths:
            # proportional width
            # table width is 100% of wrapper (which is estimated_px_width)
            # so each col is w / total * 100 % ? 
            # Or just set width in px?
            px = w * 7.5
            html.append(f'<col style="width: {px}px;">')
        html.append('</colgroup>')
        
        # Rows
        for r in range(min_row, max_row + 1):
            html.append('<tr>')
            for c in range(min_col, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                
                # Check merge
                is_merged = False
                is_top_left = False
                colspan = 1
                rowspan = 1
                
                # Optimizing merge check: use sheet.merged_cells.ranges
                # Still O(ranges) per cell. 
                
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            is_top_left = True
                            rowspan = merged_range.max_row - merged_range.min_row + 1
                            colspan = merged_range.max_col - merged_range.min_col + 1
                        break
                
                if is_merged and not is_top_left:
                    continue
                
                style = self._get_cell_style(cell)
                value = cell.value if cell.value is not None else ""
                
                attr = f'style="{style}"'
                if rowspan > 1: attr += f' rowspan="{rowspan}"'
                if colspan > 1: attr += f' colspan="{colspan}"'
                
                html.append(f'<td {attr}>{value}</td>')
            html.append('</tr>')
            
        html.append('</table>')
        
        # Insert QR Code if provided
        if qr_code_base64:
            html.append(f'<img src="{qr_code_base64}" class="qr-code">')
            
        html.append('</div>') # wrapper
        html.append('</body></html>')
        
        return "".join(html)

    def _get_cell_style(self, cell):
        """
        Extracts CSS styles from an openpyxl cell.
        """
        styles = []
        
        # Font
        if cell.font:
            if cell.font.b: styles.append("font-weight: bold;")
            if cell.font.i: styles.append("font-style: italic;")
            if cell.font.sz: styles.append(f"font-size: {cell.font.sz}pt;")
            # if cell.font.name: styles.append(f"font-family: '{cell.font.name}', serif;") 
            # Force Times New Roman as per requirement, or respect excel?
            # Requirement: "O'zbek tilidagi lotin va kirill harflarini to'liq qo'llab-quvvatlash uchun TrueType (.ttf) shriftlaridan (masalan, Times New Roman yoki Arial) foydalan."
            # So we stick to Times New Roman generally, but can fallback to cell font if needed.
        
        # Alignment
        if cell.alignment:
            if cell.alignment.horizontal: styles.append(f"text-align: {cell.alignment.horizontal};")
            if cell.alignment.vertical: 
                valign = cell.alignment.vertical
                if valign == 'center': valign = 'middle'
                styles.append(f"vertical-align: {valign};")
            if cell.alignment.wrap_text: styles.append("white-space: normal;")
            else: styles.append("white-space: nowrap;")
            
        # Border
        # Simple border mapping
        if cell.border:
            # We can inspect left, right, top, bottom.
            # If any side has style, add border.
            # Simplified:
            if cell.border.left.style: styles.append("border-left: 1px solid black;")
            if cell.border.right.style: styles.append("border-right: 1px solid black;")
            if cell.border.top.style: styles.append("border-top: 1px solid black;")
            if cell.border.bottom.style: styles.append("border-bottom: 1px solid black;")

        # Background (Fill)
        # if cell.fill and cell.fill.patternType == 'solid':
        #    fg = cell.fill.fgColor.rgb
        #    # RGB is often AARRGGBB, need to handle
        #    pass 

        return " ".join(styles)
