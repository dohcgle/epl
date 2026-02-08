
import openpyxl
import os

files = [
    '/home/ulugbek/Projects/epl/material/TURAYEV ISROIL BURIYEVICH.xlsx'
]

for f in files:
    print(f"--- Analyzing {os.path.basename(f)} ---")
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        
        target_sheet = None
        for sheet in wb.sheetnames:
            if 'umumiy' in sheet.lower() or 'umummiy' in sheet.lower():
                target_sheet = sheet
                break
        
        if target_sheet:
            print(f"Reading sheet: {target_sheet} (All rows, col A-B)")
            ws = wb[target_sheet]
            
            # Print label and value pairs to identify form fields
            fields = []
            for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
                if row[0]: # If label exists
                    label = str(row[0]).replace('\n', ' ').strip()
                    value = row[1]
                    fields.append((label, value))
                    print(f"Label: {label} | Value: {value}")
            
        else:
            print("Target sheet 'umumiy_malumot' not found.")
            
    except Exception as e:
        print(f"Error: {e}")
